from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List
from datetime import datetime
import io

from app.models.contract import Contract

class ExcelExporter:
    """Excel 导出工具"""
    
    # 定义导出的字段和表头
    EXPORT_FIELDS = [
        ('teacher_code', '员工工号'),
        ('name', '姓名'),
        ('department', '部门'),
        ('position', '职务'),
        ('gender', '性别'),
        ('age', '年龄'),
        ('nation', '民族'),
        ('political_status', '政治面貌'),
        ('id_number', '身份证号码'),
        ('birthplace', '籍贯'),
        ('entry_date', '入职日期'),
        ('regular_date', '转正日期'),
        ('contract_start', '合同开始日'),
        ('contract_end', '合同到期日'),
        ('job_status', '在职状态'),
        ('resign_date', '离职日期'),
        ('phone_number', '电话号码'),
        ('education', '最高学历'),
        ('graduation_school', '毕业院校'),
        ('diploma_no', '毕业证号'),
        ('graduation_date', '毕业时间'),
        ('major', '专业'),
        ('degree', '学位'),
        ('degree_no', '学位证号'),
        ('teacher_cert_type', '教师资格种类'),
        ('teacher_cert_no', '教师资格证号'),
        ('title_rank', '职称等级'),
        ('title_cert_no', '职称证号'),
        ('title_cert_date', '职称取证时间'),
        ('start_work_date', '参加工作时间'),
        ('teaching_years', '教龄'),
        ('psychology_cert', '心理证'),
        ('certificate_type', '持证类别'),
        ('teaching_grade', '任教年级'),
        ('teaching_subject', '任教学科'),
        ('last_work', '上一份工作经历'),
        ('address', '家庭住址'),
        ('emergency_contact', '紧急联系人'),
        ('emergency_phone', '联系电话'),
        ('mandarin_level', '普通话等级'),
        ('remarks', '备注'),
        ('ocr_confidence', 'OCR 平均置信度'),
        ('created_at', '创建时间'),
        ('updated_at', '更新时间'),
    ]
    
    @staticmethod
    def export_contracts(contracts: List[Contract]) -> bytes:
        """
        导出合同列表为 Excel 文件
        返回：Excel 文件的字节数据
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "教师合同信息"
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        for col_idx, (_, header) in enumerate(ExcelExporter.EXPORT_FIELDS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 写入数据
        for row_idx, contract in enumerate(contracts, start=2):
            for col_idx, (field, _) in enumerate(ExcelExporter.EXPORT_FIELDS, start=1):
                value = getattr(contract, field, '')
                
                # 格式化日期
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d')
                elif field == 'ocr_confidence' and value not in (None, ''):
                    value = f"{round(float(value) * 100, 2)}%"
                
                # 处理 None 值
                if value is None:
                    value = ''
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 自动调整列宽
        for col_idx, (_, header) in enumerate(ExcelExporter.EXPORT_FIELDS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = len(header) * 2 + 5
        
        # 添加水印信息
        ws.cell(row=len(contracts) + 3, column=1, value=f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=len(contracts) + 4, column=1, value="教师合同管理系统")
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()

    @staticmethod
    def generate_template() -> bytes:
        """
        生成导入模板
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "导入模板"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, (_, header) in enumerate(ExcelExporter.EXPORT_FIELDS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            ws.column_dimensions[cell.column_letter].width = len(header) * 2 + 5

        # 添加提示信息
        ws.cell(row=3, column=1, value="提示：请按照表头填写信息，日期格式为 YYYY-MM-DD，教龄自动计算可留空。")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(ExcelExporter.EXPORT_FIELDS))
        ws.cell(row=3, column=1).alignment = Alignment(horizontal="left")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

exporter = ExcelExporter()

