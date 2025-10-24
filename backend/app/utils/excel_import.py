import io
from typing import List, Tuple, Dict, Any
from datetime import datetime, date, timedelta

from openpyxl import load_workbook

from app.utils.excel_export import ExcelExporter


DATE_FIELDS = {
    'entry_date',
    'regular_date',
    'contract_start',
    'contract_end',
    'resign_date',
    'graduation_date',
    'title_cert_date',
    'start_work_date',
}

INT_FIELDS = {'age', 'teaching_years'}

MANDATORY_FIELDS = {'teacher_code', 'name'}

LABEL_TO_FIELD = {label: key for key, label in ExcelExporter.DEFAULT_EXPORT_FIELDS}
FIELD_TO_LABEL = {key: label for key, label in ExcelExporter.DEFAULT_EXPORT_FIELDS}

DATE_FORMATS = (
    '%Y-%m-%d',
    '%Y/%m/%d',
    '%Y.%m.%d',
    '%Y%m%d',
)


def parse_contracts_from_bytes(file_bytes: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    if not file_bytes:
        raise ValueError('文件为空，无法导入')

    try:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f'Excel 文件解析失败：{exc}')

    worksheet = workbook.active
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))

    column_map: Dict[int, str] = {}
    for idx, label in enumerate(header_row, start=1):
        if label is None:
            continue
        label_str = str(label).strip()
        field = LABEL_TO_FIELD.get(label_str)
        if field:
            column_map[idx] = field

    missing_columns = {field for field in MANDATORY_FIELDS if field not in column_map.values()}
    if missing_columns:
        readable = [FIELD_TO_LABEL.get(field, field) for field in missing_columns]
        raise ValueError(f"模板缺少必要列：{'、'.join(readable)}")

    records: List[Tuple[int, Dict[str, Any]]] = []
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if _is_empty_row(row):
            continue

        record: Dict[str, Any] = {}
        for col_index, value in enumerate(row, start=1):
            field = column_map.get(col_index)
            if not field:
                continue

            try:
                cleaned = _convert_cell_value(field, value)
            except ValueError as exc:
                column_name = FIELD_TO_LABEL.get(field, field)
                raise ValueError(f'第 {row_index} 行 {column_name} 格式错误：{exc}') from exc

            if cleaned is not None:
                record[field] = cleaned

        if not any(record.get(field) for field in MANDATORY_FIELDS):
            continue

        records.append((row_index, record))

    if not records:
        raise ValueError('未在 Excel 中解析到有效数据行')

    return records


def _convert_cell_value(field: str, value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()
        if value == '':
            return None

    if field in DATE_FIELDS:
        return _parse_date(value)

    if field in INT_FIELDS:
        return _parse_int(value)

    # 对于非数值字段，保持字符串
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)

    if isinstance(value, int):
        return str(value)

    return value


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        # Excel 日期序列
        try:
            base_date = datetime(1899, 12, 30)
            offset = int(value)
            return (base_date + timedelta(days=offset)).date()
        except Exception:
            value = str(value)

    value_str = str(value).strip()
    if not value_str:
        return None

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(value_str, fmt).date()
        except ValueError:
            continue

    raise ValueError(f'无法解析日期 "{value_str}"')


def _parse_int(value: Any) -> int | None:
    if value is None:
        return None

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise ValueError(f'非整数值 {value}')

    value_str = str(value).strip()
    if value_str == '':
        return None

    try:
        return int(float(value_str))
    except ValueError as exc:
        raise ValueError(f'无法转换为整数：{value_str}') from exc


def _is_empty_row(row: Tuple[Any, ...]) -> bool:
    return all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row)


